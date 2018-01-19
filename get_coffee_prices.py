import openpyxl
from openpyxl import load_workbook
import csv
import datetime
import urllib

## this easy method to DL file uses urllib, not urllib2, but works fine
urllib.urlretrieve ("http://www.ico.org/prices/pr-market-prices.xlsx", "pr-market-prices.xlsx")

wb = load_workbook(filename = 'pr-market-prices.xlsx')
worksheet = wb['Sheet1']

### grabbed these in case we want to incorporate them into the file name, or anywhere
title = (worksheet['A1'].value)
date_range = (worksheet['A2'].value) 

with open('pr-market-prices.csv', 'wb') as csvfile:
    writer = csv.writer(csvfile, delimiter=',')
    
	## ugly, brute force method to set the first row as column headers
    first_row = ['Date', 'Composite_Indicator_Price', 
	None, 
	'Colombian_Milds--USA', 'Colombian_Milds--European', 'Colombian_Milds--Group_Weighted_Ave', 
	None,
	'Other_Milds--USA', 'Other_Milds--European', 'Other_Milds--Group_Weighted_Ave', 
	None,
	'Brazilian_Naturals--USA', 'Brazilian_Naturals--European', 'Brazilian_Naturals--Group_Weighted_Ave', 
	None,
	'Robustas--USA', 'Robustas--European', 'Robustas--Group_Weighted_Ave']
    writer.writerow(first_row)

	# loops through whole sheet, writing rows that have a date value in first column
    for row in worksheet.iter_rows():
    	row_values = []
    	if isinstance(row[0].value, datetime.datetime): 
 			for cell in row:
 				row_values.append(cell.value)
	 		writer.writerow(row_values)

###  URLS:
###  https://github.com/csvconf/data-tables.csv/issues?q=is%3Aissue+is%3Aclosed
###  http://www.ico.org/coffee_prices.asp
###  http://www.ico.org/prices/pr-market-prices.xlsx
###  https://docs.python.org/2/howto/urllib2.html 
